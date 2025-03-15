# %%
import numpy as np
import pandas as pd
import os

def batch_encode_p1000(numbers):
    """
    Encode a batch of numbers using the P1000 scheme.
    
    Parameters:
      numbers (np.ndarray): Array of numbers to encode.
      
    Returns:
      sign_tokens (np.ndarray): Array of sign tokens ('+' or '-').
      mantissa_tokens (np.ndarray): Array of mantissa tokens as zero-padded strings.
      exponent_tokens (np.ndarray): Array of exponent tokens as strings (e.g., 'E+3').
    """
    # Convert input to numpy array
    numbers = np.array(numbers, dtype=float)  # Ensure floating-point
    
    # Sign tokens: vectorized assignment based on the sign of the numbers
    sign_tokens = np.where(numbers < 0, '-', '+')
    
    # Work with absolute values
    abs_numbers = np.abs(numbers)
    
    # Handle zeros separately to avoid log10 issues
    non_zero = abs_numbers > 0
    exponent = np.zeros_like(abs_numbers, dtype=int)
    normalized = np.zeros_like(abs_numbers, dtype=float)
    
    # Compute exponent where numbers are non-zero
    exponent[non_zero] = np.floor(np.log10(abs_numbers[non_zero])).astype(int)
    
    # Compute normalized values for non-zero numbers - using float powers to avoid integer power error
    normalized[non_zero] = abs_numbers[non_zero] / np.power(10.0, exponent[non_zero])
    
    # Scale the normalized values to get mantissa in [0, 1000)
    # Multiply by 100 since normalized values are in [1, 10)
    mantissas = np.round(normalized * 100).astype(int)
    
    # Correct any mantissa that rounds to 1000
    overflow = mantissas >= 1000
    if np.any(overflow):
        mantissas[overflow] = 100
        exponent[overflow] += 1
    
    # Format mantissas as 3-digit strings
    mantissa_tokens = np.array([f"{m:03d}" for m in mantissas])
    
    # Format exponent tokens as strings (e.g., 'E+3' or 'E-2')
    exponent_tokens = np.array([f"E{'+' if e >= 0 else ''}{e}" for e in exponent])
    
    return sign_tokens, mantissa_tokens, exponent_tokens


# %%


# Assuming batch_encode_p1000 function is already defined above

def process_file_with_columns(file_path, col_sep_token="[COL_SEP]"):
    """
    Process a single CSV file containing numerical data.
    Encodes each number using the P1000 scheme and preserves column structure by
    inserting a delimiter token between columns.
    
    Returns a list of tokens for the entire file.
    """
    # Read the CSV file into a DataFrame
    df = pd.read_csv(file_path,sep=r"\s+",header=None).head(1000)
    
    # We'll build a token list row by row
    token_list = []
    
    # Iterate over each row in the DataFrame
    for idx, row in df.iterrows():
        row_tokens = []
        # Process each column in the row separately
        for col in df.columns:
            # Encode the number in this cell; each cell may be a single number.
            # If a cell contains multiple numbers, you might need to adjust this.
            number = row[col]
            # For a single number, get the tokens as a list
            sign_tokens, mantissa_tokens, exponent_tokens = batch_encode_p1000([number])
            # Combine the triplet (they are arrays with one element each)
            number_tokens = [sign_tokens[0], mantissa_tokens[0], exponent_tokens[0]]
            row_tokens.extend(number_tokens)
            # Insert column delimiter after each column (except last column)
            if col != df.columns[-1]:
                row_tokens.append(col_sep_token)
        # Optionally, you can add a row delimiter (e.g., [ROW_SEP]) if needed.
        token_list.extend(row_tokens)
        token_list.append("[ROW_SEP]")
    
    return token_list

# # Example usage for a single file:
# file_path = "../Feynman_with_units/I.6.2"  # Update the path as needed
# token_list = process_file_with_colum                                                                                                                   ns(file_path)
# print(token_list[:20])  # Print first 20 tokens for inspection
# print("Total tokens in file:", len(token_list))


# %%
folder = "Feynman_with_units"

file_list = [f for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f))]
print(file_list)

# %%
print(len(file_list))

# %%
def process_all_files_with_columns(file_list,folder_path, col_sep_token="[COL_SEP]"):
    all_file_tokens = {}
    for file_name in file_list:
        file_path = os.path.join(folder_path, file_name)
        tokens = process_file_with_columns(file_path, col_sep_token)
        all_file_tokens[file_name] = tokens
        print(f"Processed {file_name}, token sequence length: {len(tokens)}")
    return all_file_tokens

# Example usage:
folder_path = "Feynman_with_units"
all_tokens = process_all_files_with_columns(file_list,folder_path)


# %%
from tokenizers import Tokenizer

tokenizer=Tokenizer.from_file("my_tokenizer.json")
tokenized_dict={}
for key,value in all_tokens.items():
    current_token_string=" ".join(value)
    encoding_result = tokenizer.encode(current_token_string)  
    tokenized_dict[key] = encoding_result.ids


    

# %%
import json
with open("dataset_encoded.json", "w") as f:
    json.dump(tokenized_dict, f, indent=4)

# %%
from openpyxl import load_workbook
wb = load_workbook('FeynmanEquations.xlsx', data_only=False)
ws = wb.active

# Iterate through rows (we start from row 2 because we skip header)
formula_col_index = 1 
formula_list=[]
# Iterate through rows (we start from row 2 because we skip header)
for row in ws.iter_rows(min_row=2):
    # get cell in formula column
    cell = row[formula_col_index - 1]  # zero-indexed
    formula_list.append(cell.value) 
    

##Save the formula_list in a text file
filename = "filenames.txt"

with open(filename, "w") as f:
    for formula in formula_list:
        f.write(formula + "\n")

print(f"Formulas saved to {filename}")



# %%
import json

DATA_END_ID = 4  # replace with the actual token ID for [DATA_END]
combined_samples = []

# 1) Load RPN file
with open("formulas_rpn.json", "r") as f:
    rpn_data = json.load(f)  # a list of dicts: [{"id": "I.6.2a", "rpn": [...]}, ...]

# 2) Load numeric-data file
with open("dataset_encoded.json", "r") as f:
    data_dict = json.load(f) # a dict with keys like "I.6.2a" : [ data tokens ], etc.

# 3) Combine data & RPN for each entry
for item in rpn_data:
    formula_id = item["id"]
    rpn_tokens = item["rpn"]

    # Find matching data tokens (same ID) in data_dict
    if formula_id in data_dict:
        data_tokens = data_dict[formula_id]
        combined = data_tokens + [DATA_END_ID] + rpn_tokens
        combined_samples.append(combined)
    else:
        print(f"Warning: {formula_id} not found in dataset_encoded.json")

# 'combined_samples' now contains your full sequences, each with data followed by [DATA_END] and formula RPN tokens.
with open("combined_samples.json", "w") as f:
    json.dump(combined_samples, f, indent=4)


# %%
import json
import random
from typing import List, Tuple
import numpy as np

import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import Dataset, DataLoader, random_split
from torch.optim.lr_scheduler import LambdaLR
###############################################################################
# 1. Dataset and Collation
###############################################################################
class TokenSequenceDataset(Dataset):
    """
    A dataset for pre-tokenized sequences stored in memory. 
    Each sequence is a list of integer token IDs.

    This class returns (input_seq, target_seq), where:
      - input_seq = sequence[:-1]
      - target_seq = sequence[1:]
    """
    def __init__(self, sequences: List[List[int]]):
        super().__init__()
        self.sequences = sequences

    def __len__(self):
        return len(self.sequences)

    def __getitem__(self, idx) -> Tuple[List[int], List[int]]:
        seq = self.sequences[idx]
        # We shift to create input/target pairs for next-token prediction
        input_seq = seq[:-1]
        target_seq = seq[1:]
        return input_seq, target_seq


class PadCollator:
    """
    A custom collator that pads input sequences to the same length in a batch,
    creating attention masks and ensuring alignment of input/target sequences.
    """
    def __init__(self, pad_token_id: int = 0):
        self.pad_token_id = pad_token_id

    def __call__(self, batch: List[Tuple[List[int], List[int]]]):
        # Extract all input/target pairs
        input_batch, target_batch = zip(*batch)

        max_len = max(len(seq) for seq in input_batch)

        padded_inputs = []
        padded_targets = []
        attention_masks = []

        for inp, tgt in zip(input_batch, target_batch):
            inp_len = len(inp)
            pad_len = max_len - inp_len

            # Pad inputs & targets
            padded_inp = inp + [self.pad_token_id] * pad_len
            padded_tgt = tgt + [self.pad_token_id] * pad_len

            # Create attention mask: 1 for real tokens, 0 for padded
            att_mask = [1] * inp_len + [0] * pad_len

            padded_inputs.append(padded_inp)
            padded_targets.append(padded_tgt)
            attention_masks.append(att_mask)

        # Convert to tensors
        padded_inputs = torch.tensor(padded_inputs, dtype=torch.long)
        padded_targets = torch.tensor(padded_targets, dtype=torch.long)
        attention_masks = torch.tensor(attention_masks, dtype=torch.long)

        return padded_inputs, padded_targets, attention_masks


###############################################################################
# 2. Model Definition
###############################################################################
class PositionalEncoding(nn.Module):
    """
    Standard sinusoidal positional encoding.
    """
    def __init__(self, d_model: int, dropout: float = 0.1, max_len: int = 50000):
        super().__init__()
        self.dropout = nn.Dropout(p=dropout)

        position = torch.arange(0, max_len, dtype=torch.float).unsqueeze(1)
        div_term = torch.exp(
            torch.arange(0, d_model, 2).float() * (-np.log(10000.0) / d_model)
        )
        pe = torch.zeros(max_len, d_model)
        pe[:, 0::2] = torch.sin(position * div_term)
        pe[:, 1::2] = torch.cos(position * div_term)
        pe = pe.unsqueeze(0)  # (1, max_len, d_model)
        self.register_buffer('pe', pe)

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        """
        x shape: (batch, seq_len, d_model)
        """
        seq_len = x.size(1)
        # Add positional encoding
        x = x + self.pe[:, :seq_len]
        return self.dropout(x)


class DecoderOnlyTransformer(nn.Module):
    """
    A causal (decoder-only) Transformer model for next-token prediction.
    """
    def __init__(
        self,
        vocab_size: int,
        d_model: int = 256,
        nhead: int = 4,
        num_layers: int = 10,
        dim_feedforward: int = 1024,
        dropout: float = 0.1,
    ):
        super().__init__()
        self.vocab_size = vocab_size
        self.d_model = d_model

        # Token Embedding
        self.token_emb = nn.Embedding(vocab_size, d_model)

        # Positional Encoding
        self.pos_encoder = PositionalEncoding(d_model, dropout=dropout)

        # Transformer Decoder Layers
        decoder_layer = nn.TransformerDecoderLayer(
            d_model=d_model, 
            nhead=nhead, 
            dim_feedforward=dim_feedforward,
            dropout=dropout,
            activation='relu'
        )
        self.transformer_decoder = nn.TransformerDecoder(
            decoder_layer, 
            num_layers=num_layers
        )

        # Final projection to vocabulary
        self.output_proj = nn.Linear(d_model, vocab_size)

        # Causal mask cache (for efficiency)
        self.register_buffer("mask_cache", None)

    def _generate_causal_mask(self, sz: int, device: torch.device):
        """
        Generates an upper-triangular causal mask to ensure each token 
        can only attend to preceding tokens (including itself).
        """
        if (self.mask_cache is None) or (self.mask_cache.size(0) < sz):
            mask = torch.triu(torch.ones(sz, sz, device=device), diagonal=1)
            # Convert to boolean, True means "block this position"
            mask = mask.bool()
            self.mask_cache = mask
        else:
            mask = self.mask_cache[:sz, :sz]
        return mask

    def forward(
        self,
        input_ids: torch.Tensor,
        attention_mask: torch.Tensor
    ) -> torch.Tensor:
        """
        input_ids: (batch, seq_len)
        attention_mask: (batch, seq_len) => 1 for real tokens, 0 for pads

        Returns:
          logits of shape (batch, seq_len, vocab_size)
        """
        device = input_ids.device
        batch_size, seq_len = input_ids.shape

        # Generate token embeddings
        tok_emb = self.token_emb(input_ids)  # (batch, seq_len, d_model)
        # Add positional encoding
        pos_emb = self.pos_encoder(tok_emb)

        # Prepare the causal mask
        causal_mask = self._generate_causal_mask(seq_len, device=device)  # (seq_len, seq_len)

        # We also need to expand the attention_mask to shape (batch, 1, seq_len)
        # so it can be broadcast to (batch, seq_len, seq_len).
        # We'll turn 0 => True in the mask to block those positions.
        extended_attention_mask = attention_mask.unsqueeze(1).repeat(1, seq_len, 1)
        # So the final mask used by the decoder is (batch, seq_len, seq_len)
        # with True where we want to block attention.
        combined_mask = causal_mask.unsqueeze(0) | (extended_attention_mask == 0)

        # Permute to fit PyTorch's (seq_len, batch, d_model)
        pos_emb = pos_emb.permute(1, 0, 2)  # => (seq_len, batch, d_model)

        # Decode (TransformerDecoder expects shape (seq_len, batch, d_model))
        # The "memory" here is empty because we're using a decoder-only approach.
        decoded = self.transformer_decoder(
            pos_emb,
            memory=torch.zeros(0, batch_size, self.d_model, device=device),  # dummy empty memory
            tgt_mask=combined_mask[0],  # shape (seq_len, seq_len) for a single batch? We'll do a trick below.
            # tgt_key_padding_mask=~attention_mask.bool()  # shape (batch, seq_len)
             tgt_key_padding_mask=~attention_mask.bool()
        )

        # NOTE: PyTorch's TransformerDecoder can’t directly handle a 3D mask. 
        # We used 'tgt_key_padding_mask' for pad tokens and 'tgt_mask' for causality. 
        # This approach merges them. If you have more advanced needs, you'd implement a custom layer or reshape.

        # Undo the permute: (seq_len, batch, d_model) -> (batch, seq_len, d_model)
        decoded = decoded.permute(1, 0, 2).contiguous()

        # Project to vocab
        logits = self.output_proj(decoded)  # (batch, seq_len, vocab_size)
        return logits


###############################################################################
# 3. Training / Validation / Testing
###############################################################################
def train_one_epoch(
    model: nn.Module,
    dataloader: DataLoader,
    optimizer: optim.Optimizer,
    criterion: nn.CrossEntropyLoss,
    device: torch.device
) -> Tuple[float, float]:
    """
    Train for one epoch. Returns (avg_loss, approx_token_accuracy).
    """
    model.train()
    running_loss = 0.0
    running_correct = 0
    total_tokens = 0

    for batch_idx, (inputs, targets, attention_mask) in enumerate(dataloader):
        inputs = inputs.to(device)
        targets = targets.to(device)
        attention_mask = attention_mask.to(device)

        optimizer.zero_grad()
        logits = model(inputs, attention_mask)  # (batch, seq_len, vocab_size)

        # Reshape for loss: (batch*seq_len, vocab_size) vs (batch*seq_len)
        vocab_size = logits.size(-1)
        loss = criterion(logits.view(-1, vocab_size), targets.view(-1))

        loss.backward()
        optimizer.step()

        # Accumulate stats
        running_loss += loss.item()
        # Approximate token-level accuracy
        preds = logits.argmax(dim=-1)  # (batch, seq_len)
        mask_flat = attention_mask.view(-1).bool()
        correct = (preds.view(-1)[mask_flat] == targets.view(-1)[mask_flat]).sum().item()
        count = mask_flat.sum().item()
        running_correct += correct
        total_tokens += count

    avg_loss = running_loss / len(dataloader)
    avg_acc = running_correct / total_tokens if total_tokens > 0 else 0.0
    return avg_loss, avg_acc


def validate(
    model: nn.Module,
    dataloader: DataLoader,
    criterion: nn.CrossEntropyLoss,
    device: torch.device
) -> Tuple[float, float]:
    """
    Validate the model. Returns (avg_loss, approx_token_accuracy).
    """
    model.eval()
    running_loss = 0.0
    running_correct = 0
    total_tokens = 0

    with torch.no_grad():
        for inputs, targets, attention_mask in dataloader:
            inputs = inputs.to(device)
            targets = targets.to(device)
            attention_mask = attention_mask.to(device)

            logits = model(inputs, attention_mask)
            vocab_size = logits.size(-1)
            loss = criterion(logits.view(-1, vocab_size), targets.view(-1))
            running_loss += loss.item()

            # Approximate token-level accuracy
            preds = logits.argmax(dim=-1)
            mask_flat = attention_mask.view(-1).bool()
            correct = (preds.view(-1)[mask_flat] == targets.view(-1)[mask_flat]).sum().item()
            count = mask_flat.sum().item()
            running_correct += correct
            total_tokens += count

    avg_loss = running_loss / len(dataloader)
    avg_acc = running_correct / total_tokens if total_tokens > 0 else 0.0
    return avg_loss, avg_acc


def test_sequence_accuracy(
    model: nn.Module,
    dataloader: DataLoader,
    device: torch.device
) -> float:
    """
    Computes the fraction of sequences where the entire predicted sequence
    matches the target sequence exactly. 
    """
    model.eval()
    correct_sequences = 0
    total_sequences = 0

    with torch.no_grad():
        for inputs, targets, attention_mask in dataloader:
            inputs = inputs.to(device)
            targets = targets.to(device)
            attention_mask = attention_mask.to(device)

            logits = model(inputs, attention_mask)
            preds = logits.argmax(dim=-1)  # (batch, seq_len)

            # For each sequence in the batch, compare all tokens (where attention_mask=1).
            for i in range(inputs.size(0)):
                seq_mask = attention_mask[i].bool()
                pred_seq = preds[i, seq_mask]
                tgt_seq = targets[i, seq_mask]
                total_sequences += 1
                if torch.equal(pred_seq, tgt_seq):
                    correct_sequences += 1

    return correct_sequences / total_sequences if total_sequences > 0 else 0.0


###############################################################################
# 4. Main Script
###############################################################################
def main():
    # ---------------------------
    # 4.1 Load Tokenized Data
    # ---------------------------
    json_path = "dataset_encoded.json"  # Change to your path if needed
    with open(json_path, "r") as f:
        token_data_dict = json.load(f)

    # token_data_dict is assumed to be {filename: [list_of_token_ids], ...}
    # Merge all token lists into one big list if needed, or keep them separate.
    # We'll merge them for a single dataset:
    all_sequences = []
    for seq_list in token_data_dict.values():
        # seq_list is presumably a list of ints
        # Possibly it's a list of lists if you segmented each file. 
        # If needed, adapt to your structure. 
        # We'll assume it's a single list of token IDs per entry.
        if isinstance(seq_list[0], int):
            # single sequence
            all_sequences.append(seq_list)
        else:
            # multiple sequences in a sub-list
            all_sequences.extend(seq_list)

    # Filter out any sequences shorter than 2 tokens (otherwise can't do input/target shift).
    all_sequences = [seq for seq in all_sequences if len(seq) > 1]

    MAX_CHUNK_LEN = 32  # or any reasonable max length
    chunked_sequences = []

    for seq_list in token_data_dict.values():
        # If seq_list is a single sequence of IDs, chunk it
        if isinstance(seq_list[0], int):
            seq = seq_list
            # Break the single seq into multiple chunks
            for i in range(0, len(seq), MAX_CHUNK_LEN):
                chunk = seq[i:i+MAX_CHUNK_LEN]
                # We only keep chunks that have at least 2 tokens
                if len(chunk) > 1:
                    chunked_sequences.append(chunk)
        else:
            # If seq_list is already a list of sequences, chunk each
            for seq in seq_list:
                for i in range(0, len(seq), MAX_CHUNK_LEN):
                    chunk = seq[i:i+MAX_CHUNK_LEN]
                    if len(chunk) > 1:
                        chunked_sequences.append(chunk)

    all_sequences = chunked_sequences


    # ---------------------------
    # 4.2 Split into Train/Val/Test
    # ---------------------------
    random.shuffle(all_sequences)
    dataset_size = len(all_sequences)
    train_size = int(0.8 * dataset_size)
    val_size = int(0.1 * dataset_size)
    test_size = dataset_size - train_size - val_size

    train_data, val_data, test_data = random_split(
        all_sequences,
        [train_size, val_size, test_size],
        generator=torch.Generator().manual_seed(42)
    )

    train_sequences = list(train_data)
    val_sequences = list(val_data)
    test_sequences = list(test_data)

    # ---------------------------
    # 4.3 Create Datasets & Loaders
    # ---------------------------
    pad_token_id = 0  # Adjust if your PAD token is something else
    train_dataset = TokenSequenceDataset(train_sequences)
    val_dataset = TokenSequenceDataset(val_sequences)
    test_dataset = TokenSequenceDataset(test_sequences)

    collator = PadCollator(pad_token_id=pad_token_id)
    batch_size = 256

    train_loader = DataLoader(train_dataset, batch_size=batch_size, shuffle=True, collate_fn=collator)
    val_loader = DataLoader(val_dataset, batch_size=batch_size, shuffle=False, collate_fn=collator)
    test_loader = DataLoader(test_dataset, batch_size=batch_size, shuffle=False, collate_fn=collator)

    sample_batch = next(iter(train_loader))
    print("Sample input shape:", sample_batch[0].shape)
    print("Sample target shape:", sample_batch[1].shape)
    print("Sample mask shape:", sample_batch[2].shape)
    print("Sample input:", sample_batch[0][0][:10])  # First 10 tokens of first example
    print("Sample target:", sample_batch[1][0][:10])
    print("Max token ID in input:", sample_batch[0].max().item())
    print("Max token ID in target:", sample_batch[1].max().item())
        # ---------------------------
    # 4.4 Model, Optimizer, Loss
    # ---------------------------
    # You likely know your max vocab size from your tokenizer
    # For example, if you used Byte-Level BPE or a custom mapping:
    vocab_size = 2000  # <-- Replace with your actual vocabulary size
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    model = DecoderOnlyTransformer(
        vocab_size=vocab_size,
        d_model=256,          # Hidden dimension
        nhead=8,              # Number of attention heads
        num_layers=6,         # Number of transformer decoder layers
        dim_feedforward=1024, # FFN dimension
        dropout=0.1
    )
    model.to(device)    
    optimizer = optim.AdamW(model.parameters(), lr=5e-4)
    pad_token_id=0
    criterion = nn.CrossEntropyLoss(ignore_index=pad_token_id)

    # ---------------------------
    # 4.5 Training Loop
    # ---------------------------
    epochs = 50
    best_val_loss = float('inf')

    for epoch in range(1, epochs + 1):
        # --- Train ---
        train_loss, train_acc = train_one_epoch(model, train_loader, optimizer, criterion, device)
        # --- Validate ---
        val_loss, val_acc = validate(model, val_loader, criterion, device)

        print(f"[Epoch {epoch}/{epochs}] "
              f"Train Loss: {train_loss:.4f} | Train Acc: {train_acc:.4f} || "
              f"Val Loss: {val_loss:.4f} | Val Acc: {val_acc:.4f}")

        # Checkpoint if validation loss improves
        if val_loss < best_val_loss:
            best_val_loss = val_loss
            torch.save(model.state_dict(), "transformer_model.pt")
            print("  (Best validation loss so far, saving model)")

    print("Training complete. Loading best model weights from checkpoint.")
    model.load_state_dict(torch.load("transformer_model.pt"))

    # ---------------------------
    # 4.6 Evaluate on Test Set (Sequence Accuracy)
    # ---------------------------
    seq_acc = test_sequence_accuracy(model, test_loader, device)
    print(f"Test Full-Sequence Accuracy: {seq_acc:.4f}")


if __name__ == "__main__":
    main()



