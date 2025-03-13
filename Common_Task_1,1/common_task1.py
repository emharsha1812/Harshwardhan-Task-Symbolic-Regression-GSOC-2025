# %% [markdown]
# # Using openpyxl for Direct Parsing of Formulas
# 
# Using `pd.read_csv()` and `pd.read_excel()` excludes some formulas. Specifically, 30 formulas are returned as NaN for CSV files and 3 formulas are returned as NaN for Excel files. Therefore, I have used `openpyxl` for direct parsing of the formulas.
# 
# CSV files are plain text and don't preserve cell metadata such as the actual formula text; they only store the computed values. When Excel exports to CSV, it writes out the result of the formula, not the formula itself. This is why we use libraries like `openpyxl` with the Excel format (XLSX) and set `data_only=False` to retrieve the underlying formula strings.

# %%
from openpyxl import load_workbook
wb = load_workbook('FeynmanEquations.xlsx', data_only=False)
ws = wb.active

# Iterate through rows (we start from row 2 because we skip header)
formula_col_index = 4 
formula_list=[]
# Iterate through rows (we start from row 2 because we skip header)
for row in ws.iter_rows(min_row=2):
    # get cell in formula column
    cell = row[formula_col_index - 1]  # zero-indexed
    formula_list.append(cell.value) 
    

##Save the formula_list in a text file
filename = "formulas.txt"

with open(filename, "w") as f:
    for formula in formula_list:
        f.write(formula + "\n")

print(f"Formulas saved to {filename}")


# %% [markdown]
# ## Plan for Tokenization
# 
# ### Extract All Unique Variable Names
# 1. Read the `FeynmanEquations.xlsx` file.
# 2. Collect all unique variable names from the columns `v1_name`, `v2_name`, `v3_name`, etc.
# 
# ### Extract All Unique Operators and Functions
# 1. Parse the formula column.
# 2. Extract mathematical operators (`+`, `-`, `*`, `/`, `**`) and functions (`sin`, `cos`, `exp`, etc.).
# 
# ### Merge into a Single Set (Vocabulary)
# 1. Combine the extracted variable names and operators/functions into a single unique set.
# 2. This set forms our tokenization vocabulary.
# 
# ### Tokenize Each Formula Using This Vocabulary
# 1. Convert each formula into a sequence of tokens using the extracted vocabulary.
# 

# %%
## Extract all Uique variable names
import pandas as pd
import sympy
import re

df=pd.read_excel('FeynmanEquations.xlsx')
variable_columns=[col for col in df.columns if re.match(r'v\d+_name',col)]


# print(variable_columns)

unique_variables=set()
for col in variable_columns:
    unique_variables.update(df[col].dropna().astype(str).unique())

print(unique_variables)



# %%
import re

mathematical_symbols = ['+', '-', '*', '/', '**', 'exp', 'sqrt', 'pi', 
                        'sin', 'cos', 'ln', 'Int', 'tanh', 'log', 
                        'arcsin', 'arctan', 'arccos']

mantissa_tokens = [f"{i:03d}" for i in range(0, 1000)]
exponent_tokens = [f"E-{i}" for i in range(11)] + [f"E+{i}" for i in range(11)]

TOKEN_DICT = {
    "<PAD>": 0,
    "<UNK>": 1,
    "[COL_SEP]": 2,
    "[ROW_SEP]": 3
}

# a-z
for i, letter in enumerate("abcdefghijklmnopqrstuvwxyz", start=len(TOKEN_DICT)):
    TOKEN_DICT[letter] = i

# A-Z
for i, letter in enumerate("ABCDEFGHIJKLMNOPQRSTUVWXYZ", start=len(TOKEN_DICT)):
    TOKEN_DICT[letter] = i

# digits 0-9
for i in range(10):
    TOKEN_DICT[str(i)] = len(TOKEN_DICT)

# A_0..Z_10 and A0..Z10
for letter in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
    for num in range(11):
        TOKEN_DICT[f"{letter}_{num}"] = len(TOKEN_DICT)
        TOKEN_DICT[f"{letter}{num}"]  = len(TOKEN_DICT)

# Add all mathematical symbols
base_index = max(TOKEN_DICT.values()) + 1
for i, symbol in enumerate(mathematical_symbols):
    TOKEN_DICT[symbol] = base_index + i


current_index = max(TOKEN_DICT.values()) + 1
for var in unique_variables:
    TOKEN_DICT[var] = current_index
    current_index += 1

# Add mantissa tokens
current_index = max(TOKEN_DICT.values()) + 1
for i, token in enumerate(mantissa_tokens, start=current_index):
    TOKEN_DICT[token] = i

# Add exponent tokens
current_index = max(TOKEN_DICT.values()) + 1
for i, token in enumerate(exponent_tokens, start=current_index):
    TOKEN_DICT[token] = i


def extract_equation_tokens(equation: str):
    pattern = r"(?:\*\*|[+\-*/=()^]|[A-Za-z_]+|\d+(?:\.\d+)?)"
    return re.findall(pattern, equation)

new_tokens = set()
for eq in formula_list:
    tokens_in_eq = extract_equation_tokens(eq)
    new_tokens.update(tokens_in_eq)

current_index = max(TOKEN_DICT.values()) + 1
for token in new_tokens:
    if token not in TOKEN_DICT:
        TOKEN_DICT[token] = current_index
        current_index += 1


# %%
##  Use TOKEN_DICT as a “Seed” for a Subword Tokenize
from tokenizers import Tokenizer, trainers, models, pre_tokenizers


# Step 1: Convert your existing TOKEN_DICT to special tokens
special_tokens = list(TOKEN_DICT.keys())  # e.g. ["<PAD>", "<UNK>", "+", "-", "*", ...]


# Step 2: Initialize a subword tokenizer (e.g. BPE)
tokenizer = Tokenizer(models.BPE(unk_token="<UNK>"))
tokenizer.pre_tokenizer = pre_tokenizers.Whitespace()

# Step 3: Prepare trainer
trainer = trainers.BpeTrainer(
    vocab_size=2000,  # or however large
    special_tokens=special_tokens
)



# Step 4: Provide your corpus in lines, ideally pre-tokenized or raw text
files = ["formulas.txt"]


tokenizer.train(files, trainer)

# Step 6: Test the resulting tokenizer
encoded = tokenizer.encode("E = sin(pi*x) + theta_12")
print("Tokens:", encoded.tokens)

# %%
tokenizer.save("my_tokenizer.json")

# %% [markdown]
# ### Tokenizing the formulaes using the tokenizer
# 

# %%
from tokenizers import Tokenizer

tokenizer=Tokenizer.from_file("my_tokenizer.json")

encoded_lines=[]
with open("formulas.txt",'r') as f:
    for line in f:
        formula=line.strip()
        encoded=tokenizer.encode(formula)
        encoded_lines.append(encoded)

for enc in encoded_lines:
    print("Tokens:", enc.tokens)
    print("Token IDs:", enc.ids)
    print()



# %%
# Now we will decode some tokens and compare them with the original formulas.
# If the decoded tokens and the original formulas are semantically similar, we can conclude that the tokenization process is effective.

# %%
enc_id = encoded_lines[0]
print(enc_id.ids)
print(enc_id.tokens)
# print(enc_id.)
decoded_text = tokenizer.decode(enc_id.ids,skip_special_tokens=False)
print(decoded_text)


# %%



